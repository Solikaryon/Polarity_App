import csv
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional

# Application created by: Luis Fernando Monjaraz Briseño / Aplicación creada por: Luis Fernando Monjaraz Briseño
# This application was created for Jabil / Esta aplicación fue creada para Jabil.

try:
	# openpyxl is a lightweight dependency for creating .xlsx files / openpyxl es una dependencia ligera para crear archivos .xlsx
	from openpyxl import Workbook
	from openpyxl.drawing.image import Image as XLImage
	from openpyxl.utils import get_column_letter
	from openpyxl.utils.units import pixels_to_EMU
	from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
	from openpyxl.drawing.xdr import XDRPositiveSize2D
except ImportError:  # pragma: no cover - handled at runtime via GUI
	Workbook = None  # type: ignore
	XLImage = None  # type: ignore
	get_column_letter = None  # type: ignore
	pixels_to_EMU = None  # type: ignore
	AnchorMarker = None  # type: ignore
	OneCellAnchor = None  # type: ignore
	XDRPositiveSize2D = None  # type: ignore

import tkinter as tk
from tkinter import filedialog, messagebox, ttk


OUTPUT_COLUMNS = [
	"Modulo",
	"Slot",
	"Part Number",
	"Package",
	"Program's Polarities",
	"Package's polarities",
	"Name of shape",
	"Coments or presentations changes",
]


REQUIRED_FEEDER_COLUMNS = [
	# Encabezados usados por la macro original
	"ModuleNumber",
	"SideNo",
	"Location",
	"PartNumber",
	"PartComment",
	"PackageName",
	"PartShapeName",
	"FeederType",
	"TapeWidth",
	"FeedPitch",
	"QTY",
]


class PolarityAppError(Exception):
	"""Application domain error (used to show clear messages in the UI)."""


@dataclass
class ProcessResult:
	output_path: Path
	rows_written: int
	images_found: int


class PolarityDataProcessor:
	"""Encapsulates the data reading logic and Excel generation."""

	def __init__(self, root_dir: Path) -> None:
		self.root_dir = root_dir
		self.data_dir = self.root_dir / "DATA"
		self.images_dir = self.root_dir / "IMAGES"
		# Job data for print headers / Datos del job para encabezados de impresión
		self.job_name: str = ""
		self.revision: str = ""
		self.side: str = ""
		# Line name extracted from FeederSetup header rows / Nombre de línea extraído de las filas de encabezado de FeederSetup
		self.line_name: str = ""

	# ------------------------------------------------------------------
	# Directories and files / Directorios y archivos
	# ------------------------------------------------------------------
	def _resolve_existing_dir(self, candidate_names: List[str]) -> Optional[Path]:
		"""Return the first existing subdirectory that matches the list.

		Allows being case-insensitive tolerant (DATA vs Data, IMAGES vs Images).
		"""

		for name in candidate_names:
			candidate = self.root_dir / name
			if candidate.is_dir():
				return candidate
		return None

	def ensure_structure(self) -> None:
		"""Create DATA and IMAGES folders inside ``root_dir`` if needed."""

		try:
			self.data_dir.mkdir(parents=True, exist_ok=True)
			self.images_dir.mkdir(parents=True, exist_ok=True)
		except OSError as exc:  # Permission/path problems / Problemas de permisos/rutas
			raise PolarityAppError(f"Could not create DATA/IMAGES folders: {exc}") from exc

	def _open_csv(self, feeder_path: Path) -> Iterable[Dict[str, str]]:
		"""Open FeederSetup.csv imitating the behavior of the original Excel macro.

		The macro validates headers in rows 1 and 3 and then deletes
		rows 1 and 2, leaving row 3 as the real table header.
		Here we replicate that: we use the third row as header and the
		rest as data.

		It first tries utf-8-sig and, if it fails, latin-1 (common in Windows exports).
		"""

		if not feeder_path.is_file():
			raise PolarityAppError(f"Required file not found: {feeder_path.name}")

		last_exc: Optional[Exception] = None
		for encoding in ("utf-8-sig", "latin-1"):
			try:
				with feeder_path.open("r", newline="", encoding=encoding) as f:
					sample = f.read(4096)
					f.seek(0)
					try:
						dialect = csv.Sniffer().sniff(sample)
					except csv.Error:
						dialect = csv.excel

					reader = csv.reader(f, dialect=dialect)
					rows = list(reader)

					# There must be at least 3 rows: header 1, job data, header 2 / Debe haber al menos 3 filas: encabezado 1, datos job, encabezado 2
					if len(rows) < 3:
						raise PolarityAppError("FeederSetup.csv does not have the expected format (too few rows).")

					# Try to read the Line name from the job info rows.
					# Row 1 (index 0) contains labels; Row 2 (index 1) contains values.
					try:
						job_headers = [h.strip() for h in rows[0]]
						job_values = rows[1] if len(rows) > 1 else []
						def _norm(s):
							return (s or "").strip().lower().rstrip(":")
						idx = next((i for i, h in enumerate(job_headers) if _norm(h) == "line"), None)
						if idx is not None and idx < len(job_values):
							self.line_name = (job_values[idx] or "").strip()
						elif len(job_values) >= 8:  # fallback to H2 when available
							self.line_name = (job_values[7] or "").strip()
					except Exception:
						# If parsing fails, keep default and continue / Si falla el parseo, mantener por defecto y continuar
						pass

					# Row 3 (index 2) is the main table header / Fila 3 (índice 2) es el encabezado principal de la tabla
					headers = [h.strip() for h in rows[2]]
					self._validate_headers(headers)

					data_rows = rows[3:]
					result: List[Dict[str, str]] = []
					for r in data_rows:
						if not any(cell.strip() for cell in r):
							continue  # skip completely empty rows / saltar filas completamente vacías
						row_dict = {h: (r[i].strip() if i < len(r) else "") for i, h in enumerate(headers)}
						result.append(row_dict)

					return result
			except UnicodeError as exc:
				last_exc = exc
				continue
			if last_exc is not None:
				raise PolarityAppError(
					"FeederSetup.csv could not be read due to encoding problems (UTF-8/Latin-1)."
				) from last_exc
		return []

	def _validate_headers(self, headers: List[str]) -> None:
		missing = [h for h in REQUIRED_FEEDER_COLUMNS if h not in headers]
		if missing:
			raise PolarityAppError(
				"Required columns are missing in FeederSetup.csv: " + ", ".join(missing)
			)

	# ------------------------------------------------------------------
	# Data preparation / Preparación de datos
	# ------------------------------------------------------------------
	def _detect_aimex(self, rows: Iterable[Dict[str, str]]) -> bool:
		"""Replicate the AIMEX/NXT logic from the original macro.

		- SideNo "1" is mapped to "2" (BOT)
		- SideNo "0" is mapped to "1" (TOP)
		- If there is at least one "2" => AIMEX = True
		"""

		aimex = False
		for row in rows:
			side = (row.get("SideNo") or "").strip()
			if side == "1":
				mapped = "2"
			elif side == "0":
				mapped = "1"
			else:
				mapped = side
			if mapped == "2":
				aimex = True
		return aimex

	def _build_slot_string(self, row: Dict[str, str], aimex: bool) -> str:
		module = (row.get("ModuleNumber") or "").strip()
		side = (row.get("SideNo") or "").strip()
		location = (row.get("Location") or "").replace("'", "").strip()

		if side == "1":
			mapped_side = "2"
		elif side == "0":
			mapped_side = "1"
		else:
			mapped_side = side

		if aimex and mapped_side:
			return f"{module}-{mapped_side}-{location}".strip("-")
		return f"{module}-{location}".strip("-")

	def _collect_image_stems(self) -> List[str]:
		if not self.images_dir.exists():
			return []
		stems: List[str] = []
		for p in self.images_dir.iterdir():
			if p.is_file():
				stems.append(p.stem)
		return stems

	def _build_output_rows(
		self, feeder_rows: List[Dict[str, str]]
	) -> List[Dict[str, Optional[str]]]:
		if not feeder_rows:
			raise PolarityAppError("FeederSetup.csv does not contain data.")

		# Filter rows with QTY == 0, as in the original macro / Filtrar filas con QTY == 0, como en la macro original
		filtered_rows: List[Dict[str, str]] = []
		for row in feeder_rows:
			qty_str = (row.get("QTY") or "").strip()
			try:
				qty_val = int(qty_str) if qty_str else 0
			except ValueError:
				qty_val = 0
			if qty_val != 0:
				filtered_rows.append(row)

		if not filtered_rows:
			raise PolarityAppError("After filtering QTY=0 there are no rows left to process.")

		aimex = self._detect_aimex(filtered_rows)

		output_rows: List[Dict[str, Optional[str]]] = []
		for row in filtered_rows:
			part_number = (row.get("PartNumber") or "").strip()
			part_comment = (row.get("PartComment") or "").strip()
			package_name = (row.get("PackageName") or "").strip()
			part_shape = (row.get("PartShapeName") or "").strip()

			slot_string = self._build_slot_string(row, aimex)
			module_only = (row.get("ModuleNumber") or "").strip()
			location_only = (row.get("Location") or "").strip()

			output_row = {
				"Modulo": module_only,
				"Slot": location_only,
				"Part Number": part_number,
				"Package": package_name,
				"Program's Polarities": "",  # aquí se insertará la imagen
				"Package's polarities": "",   # reservado para llenado manual
				"Name of shape": part_shape,
				"Coments or presentations changes": part_comment,
			}
			output_rows.append(output_row)

		return output_rows

	# ------------------------------------------------------------------
	# Excel writing / Escritura de Excel
	# ------------------------------------------------------------------
	def _ensure_workbook_available(self) -> None:
		if Workbook is None:
			raise PolarityAppError(
				"The 'openpyxl' library is not installed. "
				"Install it with: pip install openpyxl"
			)

	def _ensure_image_support(self) -> None:
		"""Verify that openpyxl has image support (Pillow installed)."""

		if XLImage is None:
			raise PolarityAppError(
				"Pillow is required to insert images into Excel. "
				"Install it with: pip install pillow"
			)

	def _find_image_for_part(self, part_number: str) -> Optional[Path]:
		"""Search for an image file in IMAGES that matches the part number."""

		part_number = part_number.strip()
		if not part_number or not self.images_dir.exists():
			return None

		for ext in (".bmp", ".png", ".jpg", ".jpeg"):
			candidate = self.images_dir / f"{part_number}{ext}"
			if candidate.is_file():
				return candidate
		return None

	def _write_excel(self, rows: List[Dict[str, Optional[str]]], output_path: Optional[Path] = None) -> ProcessResult:
		self._ensure_workbook_available()
		assert Workbook is not None  # para type checkers

		wb = Workbook()
		ws = wb.active
		ws.title = "PolarityCheck"

		# --------------------------------------------------------------
		# HEADER: title / ENCABEZADO: título
		# --------------------------------------------------------------
		from openpyxl.styles import Alignment, Border, Font, Side, PatternFill

		ws.merge_cells("A1:H1")
		title_cell = ws["A1"]
		title_cell.value = "Polarity Check"
		title_cell.font = Font(bold=True, size=16)
		title_cell.alignment = Alignment(horizontal="center", vertical="center")
		ws.row_dimensions[1].height = 25

		thin = Side(border_style="thin", color="000000")
		border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
		# Soft yellow fill for passive materials (RES/CAP) / Relleno amarillo suave para materiales pasivos (RES/CAP)
		passive_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

		# --------------------------------------------------------------
		# GENERAL INFORMATION BLOCK (form) / BLOQUE DE INFORMACIÓN GENERAL (formulario)
		# --------------------------------------------------------------
		# Layout: a single column of fields, each on its own row, / Diseño: una sola columna de campos, cada uno en su propia fila,
		# with the label in A and the value space from B to H. / con la etiqueta en A y el espacio para valor de B a H.
		info_labels = [
			"Model",
			"Line",
			"Revision",
			"Machine Programmer",
			"Quality",
			"Fecha",
		]
		start_row = 3
		for i, label in enumerate(info_labels):
			row = start_row + i
			ws[f"A{row}"].value = label
			ws[f"A{row}"].font = Font(bold=True)
			ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
			ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
			for col in range(1, 9):
				ws.cell(row=row, column=col).border = border_all

		# Fill form values from header data
		info_values = {
			"Model": self.job_name or "",
			"Line": self.line_name or self.side or "",
			"Revision": self.revision or ""
		}
		for i, label in enumerate(info_labels):
			row = start_row + i
			val = info_values.get(label)
			if val:
				ws[f"B{row}"].value = val

		# --------------------------------------------------------------
		# MAIN TABLE / TABLA PRINCIPAL
		# --------------------------------------------------------------
		header_row = start_row + len(info_labels) + 1  # deja una fila en blanco
		for idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
			cell = ws.cell(row=header_row, column=idx, value=col_name)
			cell.font = Font(bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.border = border_all

		data_start_row = header_row + 1
		for r_idx, row in enumerate(rows, start=data_start_row):
			# Determine if it is a passive material based on Name of shape / Determinar si es material pasivo según el Name of shape
			shape_name = (row.get("Name of shape") or "").strip().upper()
			is_passive = False
			if "_" in shape_name:
				after_underscore = shape_name.split("_", 1)[1]
				if after_underscore:
					# Exclude connectors or other non-passives / Excluir conectores u otros no pasivos
					if not after_underscore.startswith("CON"):
						shape_type = after_underscore[0]
						is_passive = shape_type in ("C", "R")  # C = CAPACITOR, R = RESISTOR
			for c_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
				val = row.get(col_name, "")
				cell = ws.cell(row=r_idx, column=c_idx, value=val)
				# Enable wrap_text and center the content inside the cell. / Habilitar wrap_text y centrar el contenido dentro de la celda.
				cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
				cell.border = border_all
				# If it is a passive material, highlight Name of shape and comments in soft yellow / Si es material pasivo, resaltar Name of shape y comentarios en amarillo suave
				if is_passive and col_name in ("Name of shape", "Coments or presentations changes"):
					cell.fill = passive_fill

		# Ajuste de anchos: última columna más ancha
		widths = {
			"A": 12,  # Modulo
			"B": 12,  # Slot
			"C": 18,  # Part Number
			"D": 16,  # Package
			"E": 22,  # Program's Polarities (imagen)
			"F": 22,  # Package's polarities
			"G": 18,  # Name of shape
			"H": 35,  # Coments or presentations changes (más ancha)
		}
		for col, width in widths.items():
			ws.column_dimensions[col].width = width

		# Insert images in column "Program's Polarities" (column E), / Insertar imágenes en la columna "Program's Polarities" (columna E),
		# based on the value of "Part Number". / basadas en el valor de "Part Number".
		if rows:
			self._ensure_image_support()
			for offset, row in enumerate(rows):
				excel_row = data_start_row + offset
				part_number = (row.get("Part Number") or "").strip()
				image_path = self._find_image_for_part(part_number)
				if not image_path:
					continue
				img = XLImage(str(image_path))
				col_idx = 5  # columna E
				col_letter = get_column_letter(col_idx)
				col_width = ws.column_dimensions[col_letter].width or 8.43
				# Approximation: 1 column width unit ≈ 7 pixels / Aproximación: 1 unidad de ancho de columna ≈ 7 píxeles
				cell_width_px = col_width * 7
				# Ensure enough row height and convert to pixels (1 point ≈ 4/3 px) / Asegurar altura de fila suficiente y convertir a píxeles (1 punto ≈ 4/3 px)
				current_height = ws.row_dimensions[excel_row].height or 20
				ws.row_dimensions[excel_row].height = max(current_height, 70)
				row_height_pts = ws.row_dimensions[excel_row].height
				row_height_px = row_height_pts * 4 / 3

				img_width_px = img.width
				img_height_px = img.height

				# Calculate offsets to center the image inside the cell / Calcular desplazamientos para centrar la imagen dentro de la celda
				x_offset_px = max((cell_width_px - img_width_px) / 2, 0)
				y_offset_px = max((row_height_px - img_height_px) / 2, 0)

				try:
					marker = AnchorMarker(
						col=col_idx - 1,
						colOff=pixels_to_EMU(int(x_offset_px)),
						row=excel_row - 1,
						rowOff=pixels_to_EMU(int(y_offset_px)),
					)
					size = XDRPositiveSize2D(
						pixels_to_EMU(int(img_width_px)),
						pixels_to_EMU(int(img_height_px)),
					)
					img.anchor = OneCellAnchor(_from=marker, ext=size)
				except Exception:
					# If something fails with advanced anchoring, use the basic method / Si algo falla con el anclaje avanzado, usar el método básico
					img.anchor = f"E{excel_row}"

				ws.add_image(img)

		# --------------------------------------------------------------
		# FOOTER and page setup / PIE DE PÁGINA (footer) y configuración de página
		# --------------------------------------------------------------
		# Define exact print area to avoid blank pages to the right / Definir área de impresión exacta para evitar páginas en blanco a la derecha
		max_row = ws.max_row
		ws.print_area = f"A1:H{max_row}"

		# Single page width, multiple pages down / Una sola página de ancho, múltiples hacia abajo
		ws.page_setup.fitToWidth = 1
		ws.page_setup.fitToHeight = 0  # tantas páginas hacia abajo como sea necesario
		ws.page_setup.orientation = "portrait"
		try:
			ws.page_setup.fitToPage = True
		except Exception:
			# Algunas versiones antiguas de openpyxl podrían no exponer fitToPage
			pass

		# Repeat only the table header row on all pages / Repetir solo la fila de encabezados de la tabla en todas las páginas
		ws.print_title_rows = f"{header_row}:{header_row}"

		# Encabezado para páginas (JOB / Side / Rev, fecha y página)
		job = self.job_name or ""
		rev = self.revision or ""
		side = self.side or ""
		header_left = f"JOB: {job}  Side: {side}  Rev: {rev}".strip()
		date_str = datetime.now().strftime("%d/%m/%Y")
		try:
			# Use a different header for the first page: no text. / Usar encabezado distinto para la primera página: sin texto.
			ws.page_setup.differentFirst = True
			ws.firstHeader.left.text = ""
			ws.firstHeader.center.text = ""
			ws.firstHeader.right.text = ""

			ws.oddHeader.left.text = header_left
			ws.oddHeader.center.text = date_str
			ws.oddHeader.right.text = "&P de &N"
		except Exception:
			# If the openpyxl version does not support advanced headers, ignore / Si la versión de openpyxl no soporta headers avanzados, ignorar
			pass

		# Left and right footer / Footer izquierdo y derecho
		footer_left = (
			"Format: 06-ME30-ME-ALLPLANT-00599-A\n"
			"All rights reserved – Confidential document and property of Jabil"
		)
		footer_right = "Reference Document: 06-ME30-ME-ALLPLANT-03634"
		try:
			ws.oddFooter.left.text = footer_left
			ws.oddFooter.right.text = footer_right
			ws.oddFooter.center.text = ""
		except Exception:
			# If the openpyxl version does not support oddFooter, ignore silently / Si la versión de openpyxl no soporta oddFooter, ignorar silenciosamente
			pass

		timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
		if output_path is None:
			output_path = self.root_dir / f"polarity_report_{timestamp}.xlsx"

		try:
			wb.save(output_path)
		except OSError as exc:
			raise PolarityAppError(f"The Excel file could not be saved: {exc}") from exc

		images_found = len(self._collect_image_stems())
		return ProcessResult(output_path=output_path, rows_written=len(rows), images_found=images_found)

	# ------------------------------------------------------------------
	# Processing entry point / Punto de entrada de procesamiento
	# ------------------------------------------------------------------
	def run(self, output_path: Optional[Path] = None) -> ProcessResult:
		if not self.root_dir.exists() or not self.root_dir.is_dir():
			raise PolarityAppError("The selected folder does not exist or is not a valid folder.")

		# Try to get Job / Rev / Side from the root folder name / Intentar obtener Job / Rev / Side del nombre de la carpeta raíz
		folder_name = self.root_dir.name
		parts = [p.strip() for p in folder_name.split(";")]
		if len(parts) >= 3:
			self.job_name = parts[0]
			self.revision = parts[1]
			self.side = parts[2]

		# Resolve actual folders (supports DATA/Data/data and IMAGES/Images/images) / Resolver carpetas reales (soporta DATA/Data/data e IMAGES/Images/images)
		existing_data = self._resolve_existing_dir(["DATA", "Data", "data"])
		existing_images = self._resolve_existing_dir(["IMAGES", "Images", "images"])

		if existing_data is not None:
			self.data_dir = existing_data
		else:
			self.data_dir = self.root_dir / "DATA"

		if existing_images is not None:
			self.images_dir = existing_images
		else:
			self.images_dir = self.root_dir / "IMAGES"

		self.ensure_structure()

		feeder_path = self.data_dir / "FeederSetup.csv"
		feeder_rows = self._open_csv(feeder_path)
		output_rows = self._build_output_rows(list(feeder_rows))
		return self._write_excel(output_rows, output_path)


class PolarityGUI(tk.Tk):
	"""Simple graphical interface to run the process."""

	def __init__(self) -> None:
		super().__init__()
		self.title("Polarity Check")
		self.resizable(False, False)

		self.selected_dir_var = tk.StringVar(value="(no folder selected)")
		self.status_var = tk.StringVar(value="Select the root folder for the output report.")

		self._build_widgets()

	def _build_widgets(self) -> None:
		padding = {"padx": 10, "pady": 5}

		main = ttk.Frame(self)
		main.grid(row=0, column=0, sticky="nsew", **padding)

		# Folder selection / Selección de carpeta
		ttk.Label(main, text="Root folder:").grid(row=0, column=0, sticky="w")

		dir_frame = ttk.Frame(main)
		dir_frame.grid(row=1, column=0, columnspan=2, sticky="ew")
		dir_frame.columnconfigure(0, weight=1)

		dir_label = ttk.Label(dir_frame, textvariable=self.selected_dir_var, relief="sunken")
		dir_label.grid(row=0, column=0, sticky="ew", padx=(0, 5))

		ttk.Button(dir_frame, text="Browse...", command=self.on_browse).grid(row=0, column=1)

		# Run button / Botón de ejecución
		ttk.Button(main, text="Generate report", command=self.on_run).grid(
			row=2, column=0, columnspan=2, pady=(10, 5)
		)

		# Status / Estado
		status_label = ttk.Label(main, textvariable=self.status_var, wraplength=380)
		status_label.grid(row=3, column=0, columnspan=2, sticky="w", pady=(5, 0))

	# ------------------------------------------------------------------
	# Callbacks / Callbacks
	# ------------------------------------------------------------------
	def on_browse(self) -> None:
		directory = filedialog.askdirectory(title="Select root folder")
		if directory:
			self.selected_dir_var.set(directory)
			self.status_var.set("Folder selected. Ready to generate the report.")

	def on_run(self) -> None:
		root_path_str = self.selected_dir_var.get()
		if not root_path_str or root_path_str.startswith("("):
			messagebox.showwarning("Folder not selected", "Please select a root folder first.")
			return

		root_dir = Path(root_path_str)

		processor = PolarityDataProcessor(root_dir)

		# Ask the user where to save the Excel file / Preguntar al usuario dónde guardar el archivo Excel
		from datetime import datetime as _dt
		default_name = f"polarity_report_{_dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
		save_path_str = filedialog.asksaveasfilename(
			title="Save report as",
			defaultextension=".xlsx",
			filetypes=[("Archivos de Excel", "*.xlsx")],
			initialdir=str(root_dir),
			initialfile=default_name,
		)
		if not save_path_str:
			# User cancelled saving / Usuario canceló el guardado
			self.status_var.set("Saving cancelled by the user.")
			return

		output_path = Path(save_path_str)
		self.status_var.set("Processing data... Please wait.")
		self.update_idletasks()

		try:
			result = processor.run(output_path=output_path)
		except PolarityAppError as exc:
			messagebox.showerror("Processing error", str(exc))
			self.status_var.set("An error occurred. Check the message shown.")
			return
		except Exception as exc:  # Defensa adicional
			messagebox.showerror("Unexpected error", f"An unexpected error occurred: {exc}")
			self.status_var.set("An unexpected error occurred. Please contact IT.")
			return

		msg_lines = [
			f"File generated successfully:",
			str(result.output_path),
			"",
			f"Rows written: {result.rows_written}",
			f"Images detected in IMAGES: {result.images_found}",
		]
		messagebox.showinfo("Process completed", "\n".join(msg_lines))

		self.status_var.set(f"Report generated at: {result.output_path}")


def main() -> None:
	# Run as a desktop application / Ejecutar como aplicación de escritorio
	app = PolarityGUI()
	app.mainloop()


if __name__ == "__main__":
	# Allow running the script from the command line / Permite ejecutar el script desde la línea de comandos
	try:
		main()
	except KeyboardInterrupt:
		sys.exit(1)

